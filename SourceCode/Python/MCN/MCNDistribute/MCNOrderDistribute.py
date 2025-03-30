import os
import re
import openpyxl
import pandas as pd
from pub.utils.ConfigUtils import getDefaultConfigValue
from pub.utils.ExcelUtils import dfTable, setMCNOrderExcel, is_sheet_exist
from pub.utils.TimeUtils import diffBothTime, formatDateStr
from pub.utils.XlsxSaver import XlsxSaver


configPath = 'D://stocktaking/variableConfig.ini'


# 合并8仓订单详情 和 联总表匹配项目号组套要求等
def exportOrderDetailExcel(importExcelPath, exportResultPath):
    exportSheetName = '原始订单表'
    # configPath = 'D://stocktaking/variableConfig.ini'
    # 8仓订单总表
    xls = pd.ExcelFile(importExcelPath)

    # sheet表名 ：合计发货数量
    sheetNames = xls.sheet_names
    sumSheetName = None
    if '合计发货数量' in sheetNames:
        sumSheetName = '合计发货数量'
    else:
        print('订单表（' + importExcelPath + '）不包含Sheet表(订单总表)：[合计发货数量]')
        return False
    sumDataDf = pd.read_excel(importExcelPath, sheet_name=sumSheetName, dtype=str, skiprows=1)

    # 8仓订单总表列名 : '商品编号', '项目号', '商品名称', '备注贴码', '系数'
    sumColumns = sumDataDf.columns
    not_contained = [item for item in ['商品编号', '项目号', '商品名称', '备注贴码', '合计', '系数'] if item not in sumColumns]
    if len(not_contained) > 0:
        print('8仓订单总表的列名里' + not_contained.__str__() + '未匹配到')
        if '系数' not in sumColumns:
            sumDataDf['系数'] = 1
            print('系数列添加默认值：1')
        # 如果商品编号没在标题行里
        if '商品编号' not in sumColumns:
            # 就判断商品编码是否在列表标题内，并替换成商品编号
            if '商品编码' in sumColumns:
                sumDataDf = sumDataDf.rename(columns={'商品编码': '商品编号'}, inplace=False)
            else:
                print('订单Sheet总表列名里商品编码或商品编号都未匹配到，请修改完在重新执行')
                return False

    # 8个分仓sheet表名
    sheetNamesList = []
    for sheetName in sheetNames:
        orderSheetRe = getDefaultConfigValue(configPath, 'MCN', 'orderSheetRe', '.*北京|上海|广州|成都|武汉|沈阳|西安|德州.*')
        if re.match(orderSheetRe, sheetName):
            sheetNamesList.append(sheetName)
    print('共获取' + str(len(sheetNamesList)) + '个仓的数据' + sheetNamesList.__str__())

    # 把8个分仓Sheet合并成 ->订单详情总表
    orderDetailDf = pd.concat([pd.read_excel(importExcelPath, ls, dtype=str, skiprows=1) for ls in sheetNamesList])

    orderDetailColumns = orderDetailDf.columns
    # 订单详情总表列名 :采购单号,商品编号,商品名称,配送中心,原始采购数量
    orderDetailColumnsStr = getDefaultConfigValue(configPath, 'MCN', 'orderColumnsStr', '采购单号,商品编号,商品名称,配送中心,原始采购数量')
    orderDetailColumnsList = orderDetailColumnsStr.split(',')
    not_contained_orderDetail = [item for item in orderDetailColumnsList if item not in orderDetailColumns]
    if len(not_contained_orderDetail) > 0:
        print('订单详情表的列名里' + not_contained_orderDetail.__str__() + '未匹配到')
        if '原始采购数量' not in orderDetailColumns:
            if '采购数量' in orderDetailColumns:
                sumDataDf = sumDataDf.rename(columns={'采购数量': '原始采购数量'}, inplace=False)
            else:
                print('订单详情表的列名里采购数量或原始采购数量都未匹配到，请修改完在重新执行')
                return False
    # 如果两个列名都有
    if all(item in orderDetailColumns for item in ['采购数量', '原始采购数量']):
        orderDetailDf['原始采购数量'] = orderDetailDf.apply(lambda row: row['采购数量'] if not pd.isna(row['采购数量']) else row['原始采购数量'], axis=1)
        # 删除列
        orderDetailDf = orderDetailDf.drop('采购数量', axis=1)
    # 保留列
    orderDetailDf = orderDetailDf[orderDetailColumnsList]
    # 过滤
    orderDetailDf.fillna('', inplace=True)
    passRowStr = getDefaultConfigValue(configPath, 'MCN', 'passRowStr', '合计|入库交接|采购单号')
    orderDetailDf = orderDetailDf[~orderDetailDf['采购单号'].str.match(passRowStr)]
    # 联表查询
    sumDataDf['商品编号'] = sumDataDf['商品编号'].fillna("")
    sumDataDf['项目号'] = sumDataDf['项目号'].fillna("")
    dfMerge = pd.merge(orderDetailDf, sumDataDf[['项目号', '备注贴码', '商品编号', '系数']], left_on='商品编号',
                       right_on='商品编号', how='left', indicator=True)
    # 礼袋加到最后
    liDaiDf = sumDataDf[(sumDataDf['商品编号'] == '') & (sumDataDf['项目号'] != "")]
    liDaiDf['配送中心'] = '赠品'
    liDaiDf = liDaiDf.rename(columns={'合计': '原始采购数量'}, inplace=False)
    dfMerge = pd.concat([dfMerge, liDaiDf[['项目号', '配送中心', '原始采购数量']]], ignore_index=True)
    # 设置默认值
    dfMerge['系数'] = dfMerge['系数'].fillna(1)
    dfMerge['原始采购数量'] = dfMerge['原始采购数量'].fillna(0)
    dfMerge['原始采购数量'] = dfMerge['原始采购数量'].astype(int)
    # 转换数据类型
    dfMerge['订单盒数'] = dfMerge['原始采购数量'].astype(int) * dfMerge['系数'].astype(int)
    dfMerge['分配序号'] = range(1, len(dfMerge) + 1)
    # 未匹配到的商品编码
    columnsLeftOnly = dfMerge[dfMerge['_merge'] == 'left_only']
    if len(columnsLeftOnly) > 0:
        print('共有' + str(len(columnsLeftOnly)) + '条数据，在8仓与总表联查过程中匹配失败。详细信息：')
        dfTable(columnsLeftOnly)
    if not os.path.exists(exportResultPath):
        dfMerge.to_excel(exportResultPath, sheet_name='原始订单表', index=False)
    else:
        with pd.ExcelWriter(exportResultPath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            dfMerge.to_excel(writer, sheet_name=exportSheetName, index=False)
    return True


# 订单分配库存
def exportDistributeResult(importInventoryExcelPath, exportResultPath):
    exportSheetName = '原始订单表'
    orderDf = pd.read_excel(exportResultPath, sheet_name=exportSheetName, dtype=str)
    orderDf = orderDf.fillna("")
    orderDf = orderDf.astype({'订单盒数': int, '系数': int, '原始采购数量': int})
    # 总计盒数 用于校对分配的结果
    orderSum = orderDf['订单盒数'].sum()
    # 库存表
    inventoryDf = pd.read_excel(importInventoryExcelPath, dtype=str)
    inventoryColumns = inventoryDf.columns
    # 更名前的列名
    beforeColumns = ['第二 项目号', '说明', '生产 日期', '有效 期至', '批次/序列', '现有 数量']
    not_contained = [item for item in beforeColumns if item not in inventoryColumns]
    if len(not_contained) > 0:
        print('库存表的列名里' + not_contained.__str__() + '未匹配到')
        return False
    # 更名后的列名
    reserve_columns = ['项目号', '商品名', '生产日期', '失效日期', '批号', '数量']
    # 更改列名
    renameDict = {k: v for k, v in zip(beforeColumns, reserve_columns)}
    inventoryDf = inventoryDf.rename(columns=renameDict, inplace=False)
    # 保留列
    inventoryDf = inventoryDf[reserve_columns]
    # 更改日期格式
    inventoryDf['生产日期'] = inventoryDf['生产日期'].fillna('').apply(lambda row: formatDateStr(row, 'YYYY-MM-DD'))
    inventoryDf['失效日期'] = inventoryDf['失效日期'].fillna('').apply(lambda row: formatDateStr(row, 'YYYY-MM-DD'))
    # 分组求和
    groupList = inventoryDf.columns.tolist()
    groupList.remove('数量')
    inventoryDf['数量'] = inventoryDf['数量'].astype(int)
    inventoryDf = inventoryDf.groupby(groupList).agg({'数量': 'sum'})
    # 重置索引
    inventoryDf = inventoryDf.reset_index()
    # 添加额外辅助列
    inventoryDf['减账'] = 0
    inventoryDf['减账后库存'] = inventoryDf['数量']
    inventoryDf['分配明细(订单号:商品编号:数量)'] = ''
    # 排序
    inventoryDf.sort_values(by=['项目号', '生产日期'], ascending=[True, False], inplace=True)

    # 创建订单结果List
    resultOrderList = []
    resultOrderColumns = orderDf.columns.tolist()
    resultOrderColumns.append('商品名')
    resultOrderColumns.append('生产日期')
    resultOrderColumns.append('失效日期')
    resultOrderColumns.append('批号')
    resultOrderColumns.append('分配数量')
    resultOrderList.append(resultOrderColumns)

    # 开始分配
    for index, orderRow in orderDf.iterrows():
        inventoryDfEleCopy = inventoryDf.loc[inventoryDf['项目号'] == orderRow['项目号']]
        # 同一个单号不超过2个批号的标记
        tag = 2
        msg = '库存只有两个批号，第二个批号库存不足'
        for indexI, inventoryRow in inventoryDfEleCopy.iterrows():
            if tag <= 0:
                msg = '库存有2个以上的批号，但分配的批号已经超过2个了'
                print(msg)
                break
            resultOrderRowList = orderRow.tolist()
            over = False
            # 系数
            coe = int(orderRow['系数'])
            # 订单数
            orderNum = int(orderDf.loc[index, '订单盒数'])
            # 库存数
            inventoryNum = int(inventoryRow['减账后库存'])
            # 套数
            tao = None
            # 减值
            minus = None
            # 库存余量
            remainder = None
            # 如果库存大于订单数
            if inventoryNum >= orderNum:
                # 库存余数
                remainder = inventoryNum - orderNum
                # 减值
                minus = orderNum
                # 套数
                tao = orderNum // coe
                over = True
            # 单一日期库存不足
            else:
                # 套数
                tao = inventoryNum // coe
                # 减值
                minus = tao * coe
                # 库存余数
                remainder = inventoryNum - minus
            if tao is not None:
                if tao > 0:
                    # 库存减账数量
                    inventoryDf.loc[indexI, '减账后库存'] = remainder
                    inventoryDf.loc[indexI, '减账'] += minus
                    oStr = orderDf.loc[index, '采购单号'] + ':' + orderDf.loc[index, '商品编号'] + ':' + str(minus) + '   '
                    inventoryDf.loc[indexI, '分配明细(订单号:商品编号:数量)'] += oStr
                    # 订单余数
                    orderDf.loc[index, '订单盒数'] = orderNum - minus
                    # orderRow['订单盒数'] = orderNum - minus
                    resultOrderRowList.append(inventoryRow['商品名'])
                    resultOrderRowList.append(inventoryRow['生产日期'])
                    resultOrderRowList.append(inventoryRow['失效日期'])
                    resultOrderRowList.append(inventoryRow['批号'])
                    resultOrderRowList.append(minus)
                    resultOrderList.append(resultOrderRowList)
                    tag -= 1
                    if over:
                        break
        # 总库存不够
        allDifference = orderDf.loc[index, '订单盒数']
        if orderDf['订单盒数'].iloc[index] > 0:
            resultOrderRowList = orderRow.tolist()
            if tag == 0:
                # 一个订单下超过2个批号的提示
                resultOrderRowList.append(msg)
                resultOrderRowList.append(msg)
                resultOrderRowList.append(msg)
                resultOrderRowList.append(msg)
            else:
                resultOrderRowList.append('库存批号就只有一个，此批号库存不足')
                resultOrderRowList.append('库存批号就只有一个，此批号库存不足')
                resultOrderRowList.append('库存批号就只有一个，此批号库存不足')
                resultOrderRowList.append('库存批号就只有一个，此批号库存不足')
            resultOrderRowList.append(allDifference)
            resultOrderList.append(resultOrderRowList)

    resultOrderDf = pd.DataFrame(resultOrderList[1:], columns=resultOrderList[0])
    resultOrderDf['保质期（天）'] = resultOrderDf.fillna('').apply(
        lambda row: diffBothTime(row['生产日期'], row['失效日期']), axis=1)
    orderAllocationSum = resultOrderDf['分配数量'].sum()

    # 判断分配前后总数是否一致
    if orderSum != orderAllocationSum:
        print('有Bug，分配的总数和订单的总数不一致')

    # 用于打印的Sheet表
    printOrderDf = resultOrderDf[
        ['采购单号', '商品编号', '商品名称', '配送中心', '生产日期', '保质期（天）', '分配数量', '备注贴码']]
    printOrderDf['手动录入箱规'] = ''
    printOrderDf = printOrderDf.rename(columns={'备注贴码': '组套要求', '分配数量': '采购数量'}, inplace=False)
    # 输出
    printOrderDfSheet = '打印订单表'
    printBoxDf = printOrderDf[printOrderDf['配送中心'] != '赠品']
    printBox = '外箱签数据源'
    with pd.ExcelWriter(exportResultPath, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        resultOrderDf.to_excel(writer, sheet_name='订单表', index=False)
        inventoryDf.to_excel(writer, sheet_name='库存表', index=False)
        printOrderDf.to_excel(writer, sheet_name=printOrderDfSheet, index=False)
        printBoxDf.to_excel(writer, sheet_name=printBox, index=False)

    # # 格式化表格
    df = pd.read_excel(exportResultPath, sheet_name=printOrderDfSheet, engine='openpyxl', dtype=str)
    xlsx = XlsxSaver(df, exportResultPath, printOrderDfSheet)  # 初始化一个对象, 设定保存后的文件名和表名
    xlsx.set_width('采购单号', 14)  # 手动指定某列列宽
    xlsx.set_width('商品编号', 16)  # 手动指定某列列宽
    xlsx.set_width('商品名称', 50)  # 手动指定某列列宽
    xlsx.set_width('配送中心', 13)  # 手动指定某列列宽
    xlsx.set_width('生产日期', 15)  # 手动指定某列列宽
    xlsx.set_width('保质期（天）', 10)  # 手动指定某列列宽
    xlsx.set_width('采购数量', 10)  # 手动指定某列列宽
    xlsx.set_width('手动录入箱规', 8)  # 手动指定某列列宽
    xlsx.set_width('组套要求', 30)  # 手动指定某列列宽
    xlsx.set_color_alignment_font_border('5B9BD5', 'FFFFFF', None, 11, 11, '微软雅黑')
    xlsx.save()

    wb = openpyxl.load_workbook(exportResultPath)
    ws = wb[printOrderDfSheet]  # 或者使用 wb.get_sheet_by_name('Sheet1')

    # 设置采购数量数据数字
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=7, max_col=7):
        for cell in row:
            # 修改单元格的内容，不改变其他格式
            if cell.value is not None or cell.value != '':
                cell.value = int(cell.value)  # 你可以根据需要修改这部分内容
            else:
                cell.value = 0

    # 保存文件，这会保留原有的格式不变
    wb.save(exportResultPath)

    # 设置打印属性
    headerTitle = '入库交接单'
    setMCNOrderExcel(exportResultPath, printOrderDfSheet, headerTitle)

    return True
