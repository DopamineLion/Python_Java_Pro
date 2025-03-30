from types import NoneType

import openpyxl as pyxl
import pandas as pd
from tabulate import tabulate
# from openpyxl.reader.excel import load_workbook
from win32com.client.gencache import EnsureDispatch
from win32com.client import constants
from styleframe import StyleFrame, Styler, utils
from spire.xls import *
from spire.xls.common import *

# 使用'openpyxl'引擎，并设置mode为'a'以追加模式打开文件，if_sheet_exists='overlay'表示如果sheet已存在则覆盖
# with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
#     df_sheet1.to_excel(writer, sheet_name='Sheet1', index=False)


# 格式化DataFrame
def dfTable(dataFrame):
    from prettytable import PrettyTable
    # 创建 PrettyTable 对象
    table = PrettyTable()

    # 将索引名添加到 PrettyTable
    table.field_names = ["Index"] + list(dataFrame.columns)

    # 逐行添加数据，包括索引
    for index, row in dataFrame.iterrows():
        table.add_row([index] + list(row))

    # 打印格式化后的表格
    print(table)


def tableList(list):
    print(tabulate(list))

def is_sheet_exist(filename, sheetname):
    xls = pd.ExcelFile(filename)
    return sheetname in xls.sheet_names

# 执行打印
def print_file(file, sheetName, printerName):
    excel = EnsureDispatch("Excel.Application") # 打开Excel程序
    wb = excel.Workbooks.Open(file) # 打开Excel工作簿
    # wb.Worksheets(1).Select()  # 选择第一个工作表
    sht = wb.Sheets(sheetName)  # 指定工作表
    # 打印工作表 参数参考 https://learn.microsoft.com/zh-cn/office/vba/api/excel.worksheet.printout
    sht.PrintOut(ActivePrinter=printerName)
    wb.Close(constants.xlDoNotSaveChanges)  # （不保存）关闭工作簿
    excel.Quit()  # 退出Excel程序


# Excel美化
# 参考自：https://www.cnblogs.com/wang_yb/p/18070891
def beautifulExccel(excelFile, sheetName):
    dfAllBefore = pd.read_excel(excelFile, engine='openpyxl')
    df = pd.read_excel(excelFile, sheet_name=sheetName)
    style = Styler(shrink_to_fit=True)
    sf = StyleFrame(df, styler_obj=style)
    # sf.set_column_width_dict(
    #     {
    #         "库位": 11,
    #         "项目号": 11,
    #         "品名": 40,
    #         "入库批号": 16,
    #         "数量": 7,
    #     }
    # )
    header_style = Styler(
        bg_color="#5B9BD5",
        bold=True,
        font_size=11,
        horizontal_alignment=utils.horizontal_alignments.center,
        vertical_alignment=utils.vertical_alignments.center,
    )
    content_style = Styler(
        shrink_to_fit=True,
        font_size=11,
        bold=True,
        bg_color="#FFFFFF",
        horizontal_alignment=utils.horizontal_alignments.center,
    )
    row_style = Styler(
        bg_color="#DDEBF7",
        shrink_to_fit=True,
        bold=True,
        font_size=11,
        horizontal_alignment=utils.horizontal_alignments.center,
    )
    # 计算要设置背景色的行索引
    indexes = list(range(1, len(sf), 2))
    sf.apply_column_style(sf.columns, content_style)
    sf.apply_style_by_indexes(indexes, styler_obj=row_style)
    sf.apply_headers_style(header_style)
    writer = sf.to_excel(excelFile, sheet_name=sheetName)
    dfAllBefore.to_excel(writer, sheet_name='AllData', index=False)
    writer.close()


# 使用Spire.XLS设置打印属性
# 参考自 https://www.e-iceblue.cn/xls_python_document_operation/python-set-page-setup-options-in-excel.html
def setPrintOptionForSpireXLS(tagFileName, tagSheetName, tagHeaderTitle):
    # 创建 Workbook 对象
    workbook = Workbook()
    # 加载 Excel 文件
    workbook.LoadFromFile(tagFileName, engine='openpyxl')
    # 获取第一个工作表
    sheet = workbook.Worksheets[tagSheetName]
    sheet.Activate()

    # 获取工作表的 PageSetup 对象
    pageSetup = sheet.PageSetup
    # 设置工作表的上、下、左、右、页眉、页脚页边距（单位为英寸，1英寸 = 2.54厘米）
    pageSetup.TopMargin = 0.5
    pageSetup.BottomMargin = 0.4
    pageSetup.LeftMargin = 0.5
    pageSetup.RightMargin = 0.5
    pageSetup.HeaderMarginInch = 0.2
    pageSetup.FooterMarginInch = 0.2
    # 将纸张大小设置为 A4
    pageSetup.PaperSize = PaperSizeType.PaperA4
    # 将工作表内容在垂直方向上调整到一页（即所有行都适应到一页）
    pageSetup.FitToPagesTall = 0
    # 将工作表内容在水平方向上调整到一页（即所有列都适应到一页）
    pageSetup.FitToPagesWide = 1
    # 设置左侧页眉内容为"文本页眉页脚"，使用字体"宋体"，大小14
    pageSetup.CenterHeader = "&\"宋体\"&14 " + tagHeaderTitle
    # 设置中间页脚内容为当前页码
    pageSetup.CenterFooter = "第 &P 页, 共 &N 页"
    # 设置右侧页脚内容为当前日期
    pageSetup.RightFooter = "&D &T"
    # pageSetup.FirstHeaderString = '1:1'
    pageSetup.PrintTitleRows = '1:1'

    # 设置行高
    # sheet.SetRowHeight(1, 15)
    # 设置列宽
    # sheet.SetColumnWidth(1, 13)
    # sheet.SetColumnWidth(2, 14)
    # sheet.SetColumnWidth(3, 45)
    # sheet.SetColumnWidth(4, 19)
    # sheet.SetColumnWidth(5, 9)
    # 获取工作表中的特定单元格范围
    # range = sheet.Range["A1:E14"]
    # 或者获取工作表中已使用的单元格范围
    range = sheet.AllocatedRange
    # 字体 字号 加粗 （Arial使用后导致预览中的Excel单元格内容换行出问题）
    # range.Style.Font.FontName = 'Arial'
    range.Style.Font.FontName = '宋体'
    range.Style.Font.IsBold = True
    range.Style.Font.Size = 13
    # indexes = list(range(1, sheet.LastRow, 2))
    # 垂直居中
    range.Style.VerticalAlignment = VerticalAlignType.Center
    # 水平居中
    range.Style.HorizontalAlignment = HorizontalAlignType.Center
    # 自动调整单元格范围内所有列的列宽
    range.AutoFitColumns()
    # 设置第三列的列宽
    sheet.SetColumnWidth(3, 48)
    # 将单元格C9中的文本设置为自动换行
    # sheet.Range["C9"].Style.WrapText = True
    range.Style.WrapText = True
    # 自动调整单元格范围内所有行的行高
    range.AutoFitRows()
    # 在已分配范围的四周添加中等虚线边框，颜色为黑色
    range.BorderAround(LineStyleType.Thin, Color.get_Black())
    # range.BorderAround(LineStyleType.MediumDashed, Color.get_Black())
    # 在已分配范围的内部添加细线边框，颜色为黑色
    range.BorderInside(LineStyleType.Thin, Color.get_Black())
    # 将修改后的工作簿保存到新文件
    workbook.SaveToFile(tagFileName, ExcelVersion.Version2016)
    workbook.Dispose()


# 使用openpyxl方式设置打印属性
# 参考自https://blog.51cto.com/u_16099227/7194197
# https://blog.csdn.net/debrnr/article/details/124067362
def setPrintOptionsForOpenpyxl(fileName, sheetName, headerTitle):
    wb = pyxl.load_workbook(filename=fileName)
    ws = wb[sheetName]
    # 其他打印设置
    ws.page_setup.paperSize = "9"  # 纸张尺寸参见上表
    ws.page_setup.orientation = "portrait"  # 设置打印方向 values=("default", "portrait", "landscape")
    ws.page_setup.pageOrder = "downThenOver"  ##页面设置->工作表->打印顺序values=("downThenOver", "overThenDown") 先列后行,先行后列
    ws.page_setup.copies = True  # 文件->打印->份数 未测试
    # 打印标题行
    ws.print_title_rows = '1:1'
    # 页边距
    ws.page_margins.left = 0.3  # 左
    ws.page_margins.right = 0.3  # 右
    ws.page_margins.top = 0.5  # 上
    ws.page_margins.bottom = 0.4  # 下
    ws.page_margins.header = 0.1  # 页眉
    ws.page_margins.footer = 0.1  # 页脚
    # 设置页眉 左中右 left center right
    ws.oddHeader.center.text = headerTitle  # 文本
    ws.oddHeader.center.size = 16  # 字号
    ws.oddHeader.center.font = "微软雅黑"  # 字体
    ws.oddHeader.center.color = "000000"  # 16进制RGB颜色 参照PS
    # 设置页脚
    ws.oddFooter.center.text = "第&[Page]页 共&[Pages]页"
    ws.oddFooter.center.size = 12
    ws.oddFooter.center.font = "微软雅黑"
    ws.oddFooter.center.color = "000000"
    ws.oddFooter.right.text = "&D &T"

    ws.sheet_properties.pageSetUpPr.fitToPage = True  # 此行必须设置
    ws.page_setup.fitToHeight = False
    wb.save(fileName)
    wb.close()


def setPrintOptionsForOpenpyxl100X100(fileName, sheetName, headerTitle):
    wb = pyxl.load_workbook(filename=fileName)
    ws = wb[sheetName]
    # 其他打印设置
    ws.page_setup.paperSize = "9"  # 纸张尺寸参见上表
    ws.page_setup.orientation = "landscape"  # 设置打印方向 values=("default", "portrait", "landscape")
    # ws.page_setup.orientation = "portrait"  # 设置打印方向 values=("default", "portrait", "landscape")
    ws.page_setup.pageOrder = "overThenDown"  ##页面设置->工作表->打印顺序values=("downThenOver", "overThenDown") 先列后行,先行后列
    ws.page_setup.copies = True  # 文件->打印->份数 未测试
    # 打印标题行
    ws.print_title_rows = '1:1'
    # 页边距
    ws.page_margins.left = 0  # 左
    ws.page_margins.right = 0  # 右
    ws.page_margins.top = 0.3  # 上
    ws.page_margins.bottom = 0.3  # 下
    ws.page_margins.header = 0  # 页眉
    ws.page_margins.footer = 0  # 页脚
    # 设置页眉 左中右 left center right
    ws.oddHeader.center.text = headerTitle  # 文本
    ws.oddHeader.center.size = 10  # 字号
    ws.oddHeader.center.font = "微软雅黑"  # 字体
    ws.oddHeader.center.color = "000000"  # 16进制RGB颜色 参照PS
    # 设置页脚
    ws.oddFooter.center.text = "第&[Page]页 共&[Pages]页"
    ws.oddFooter.center.size = 10
    ws.oddFooter.center.font = "微软雅黑"
    ws.oddFooter.center.color = "000000"
    ws.oddFooter.right.text = "&D &T"

    ws.sheet_properties.pageSetUpPr.fitToPage = True  # 此行必须设置
    ws.page_setup.fitToHeight = False
    wb.save(fileName)
    wb.close()


# import openpyxl as pyxl
# from openpyxl.worksheet.pagebreak import Break
# wb=pyxl.Workbook()
# ws=wb.active
# #源码worksheet->worksheet->Worksheet()
# #设置打印区域
# ws.print_area = 'A1:F10'
#
# #设置打印标题和打印列
# ws.print_title_rows='1:1'
# ws.print_title_cols="A:B"
# print(ws.print_titles)#只读属性
#
# #冻结窗格 冻结第一行第一列
# ws.freeze_panes = 'B2'
#
# #未测试 分页符 适用于openpyxl 3.0.4以后
# row_number=20 #需要插入分页符的行号
# next_page_horizon, next_page_vertical = ws.page_breaks # 返回2个变量 后期版本可能取消
# next_page_horizon.append(Break(row_number)) # 通过help可以查到append属性
#
# #设置打印A3横向
# ws.set_printer_settings(ws.PAPERSIZE_A3,ws.ORIENTATION_LANDSCAPE)
# """
#     #所有默认设置如下 worksheet 类属性
#     # Paper size
#     PAPERSIZE_LETTER = '1'
#     PAPERSIZE_LETTER_SMALL = '2'
#     PAPERSIZE_TABLOID = '3'
#     PAPERSIZE_LEDGER = '4'
#     PAPERSIZE_LEGAL = '5'
#     PAPERSIZE_STATEMENT = '6'
#     PAPERSIZE_EXECUTIVE = '7'
#     PAPERSIZE_A3 = '8'
#     PAPERSIZE_A4 = '9'
#     PAPERSIZE_A4_SMALL = '10'
#     PAPERSIZE_A5 = '11'
#
#     # Page orientation
#     ORIENTATION_PORTRAIT = 'portrait' #纵向
#     ORIENTATION_LANDSCAPE = 'landscape' #横向
# """
#
# #参照源码 worksheet->page->PrintPageSetup()
# #设置缩放所有列到一页,直接设置fitToWidth=True无效,需采用如下方法
# #所有列设置为一页 逆向思维,先缩放到页面 然后适合高度改为FLASE
# ws.sheet_properties.pageSetUpPr.fitToPage = True#此行必须设置
# ws.page_setup.fitToHeight = False
# #其他打印设置
# ws.page_setup.orientation = "landscape"#设置打印方向 values=("default", "portrait", "landscape")
# ws.page_setup.paperSize = "8" #纸张尺寸参见上表
# ws.page_setup.firstPageNumber = 1#页码起始页
# ws.page_setup.useFirstPageNumber = True #使用起始页 不知道啥意思 未测试
# ws.page_setup.paperHeight = 297#纸张高度
# ws.page_setup.paperWidth = 410#纸张宽度
# ws.page_setup.pageOrder = "downThenOver"##页面设置->工作表->打印顺序values=("downThenOver", "overThenDown") 先列后行,先行后列
# ws.page_setup.usePrinterDefaults = True #使用默认打印机
# ws.page_setup.blackAndWhite = True #页面设置->工作表->单色模式
# ws.page_setup.draft = True #页面设置->工作表->草稿质量
# ws.page_setup.cellComments = True #页面设置->工作表->批注和注释values=("asDisplayed", "atEnd") 如工作表所示,工作表末尾
# ws.page_setup.errors = True #页面设置->工作表->错误单元格打印为values=("displayed", "blank", "dash", "NA")  显示值,空白,--,"#N/A"
# ws.page_setup.horizontalDpi = True #页面设置->工作表->打印质量
# ws.page_setup.verticalDpi = True #页面设置->工作表->打印质量
# ws.page_setup.copies = True #文件->打印->份数 未测试
#
#
# #参照源码worksheet->header_footer->_HeaderFooterPart()
# #设置页眉 左中右 left center right
# ws.oddHeader.center.text = "XXX" #文本
# ws.oddHeader.center.size = 24 #字号
# ws.oddHeader.center.font = "微软雅黑" #字体
# ws.oddHeader.center.color = "000000" #16进制RGB颜色 参照PS
# #设置页脚
# ws.oddFooter.center.text = "第&[Page]页 共&[Pages]页"
# ws.oddFooter.center.size = 12
# ws.oddFooter.center.font = "微软雅黑"
# ws.oddFooter.center.color = "000000"
#
# #其他页眉页脚设置 未测试
# ws.differentOddEven=True #页面设置->页眉/页脚->奇偶页不同
# ws.differentFirst=True #页面设置->页眉/页脚->首页不同
# ws.scaleWithDoc=True #页面设置->页眉/页脚->随文档自动缩放
# ws.alignWithMargins=True #页面设置->页眉/页脚->与页边距对齐
# ws.evenHeader.center.text="XXX" #偶数页眉
# ws.evenFooter.center.text="XXX" #偶数页脚
# ws.firstHeader.center.text="XXX" #奇数页眉
# ws.firstFooter.center.text="XXX" #奇数页脚
#
# # 页眉页脚自动文本
# """
# Individual left/center/right header/footer part
#
# Do not use directly.
#
# Header & Footer ampersand codes:
#
# * &A   Inserts the worksheet name #工作表名
# * &B   Toggles bold #加粗
# * &D or &[Date]   Inserts the current date #日期
# * &E   Toggles double-underline #双下划线
# * &F or &[File]   Inserts the workbook name #文件名
# * &I   Toggles italic #斜体
# * &N or &[Pages]   Inserts the total page count #总页码
# * &S   Toggles strikethrough #删除线
# * &T   Inserts the current time #当前时间
# * &[Tab]   Inserts the worksheet name #当前工作表名
# * &U   Toggles underline #下划线
# * &X   Toggles superscript #上标
# * &Y   Toggles subscript #下标
# * &P or &[Page]   Inserts the current page number #当前页码
# * &P+n   Inserts the page number incremented by n #当前页码+n
# * &P-n   Inserts the page number decremented by n #当前页码-n
# * &[Path]   Inserts the workbook path #当前文件路径
# * &&   Escapes the ampersand character #转义字符和符号
# * &"fontname"   Selects the named font #选择字体名
# * &nn   Selects the specified 2-digit font point size #选择指定的两位字体点大小?
#
# Colours are in RGB Hex #颜色是十六进制RGB
# """
#
# #源码worksheet->page->PrintOptions()
# #页面设置->页边距->居中方式 水平/垂直
# ws.print_options.horizontalCentered=True
# ws.print_options.verticalCentered=True
# #未测试
# ws.print_options.headings=True #页面设置->工作表->行和列标题
# ws.print_options.gridLines=True #页面设置->工作表->网格线
# ws.print_options.gridLinesSet=True #猜不出来什么意思
#
#
# #源码worksheet->page->PageMargins()
# #未测试 页边距
# ws.page_margins.left=0.75 #左
# ws.page_margins.right=0.75 #右
# ws.page_margins.top=1 #上
# ws.page_margins.bottom=1 #下
# ws.page_margins.header=0.5 #页眉
# ws.page_margins.footer=0.5 #页脚