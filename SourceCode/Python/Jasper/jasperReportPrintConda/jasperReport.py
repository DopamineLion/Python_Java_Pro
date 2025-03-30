import logging
import os
import traceback
from configparser import ConfigParser
from openpyxl.reader.excel import load_workbook

from pub.utils.ConfigUtils import configLog
from pub.utils.ExcelManage import jianHuoJasperSourcesExcel, waiXiangJasperSourcesExcel
from pub.utils.ExcelToJson import excelToJson
from pub.utils.ExcelUtils import is_sheet_exist
from pub.utils.FileUtils import mkDir2, copyPath, FileCopy
from pub.utils.JasperUtils import processingReportJapser
from pub.utils.TimeUtils import nowTimeSecond
from pub.utils.TimerUtils import nowTime

def exportReport(inputJasperMode, jasperSourcesExcelPath, outputJsonFilePath):
    try:
        configLog('PrintJasperLog')
        config = ConfigParser()
        config.read("D:\\stocktaking\\variableConfig.ini", "UTF-8")
        # jasperSourcesExcelPath = config.get('Jasper', "jasperSourcesExcelPath")
        # sheetName = config.get('Jasper', "sheetName")
        # pdfDir = config.get('Jasper', "pdfDir")
        # jasperModeInit = config.get('Jasper', "jasperMode")
        pdfDir = outputJsonFilePath
        # 创建PDF输出路径
        # if not os.path.exists(pdfDir):
        #     mkDir2(pdfDir)

        # inputJasperMode = None
        jasperMode = None
        jasperModeInit = None
        while True:
            # inputJasperMode = input('请选择： 拣货签输入1  外箱签输入2。回车确定')
            if inputJasperMode == '1':
                jasperModeInit = '拣货签'
                jasperMode = 'jianHuoQian'
                break
            elif inputJasperMode == '2':
                jasperModeInit = '外箱签'
                jasperMode = 'waiXiangQian'
                break
            else:
                print('未输入正确的打签类型')

        # 判断本地是否存在配置文件
        reportDir = "D:\\stocktaking\\reports\\"
        if not os.path.exists(reportDir + jasperMode + '.jrxml'):
            # 复制Jasper的配置文件
            sourcesReportDir = os.path.join(os.path.abspath(os.path.dirname(__file__)), 'pub\\utils\\reports')
            copy = FileCopy()
            copy.copy_dir(sourcesReportDir, reportDir)

        # 读取Excel数据源 和 sheet表
        # existFile = os.path.exists(jasperSourcesExcelPath)
        # if not existFile:
        #     print('配置文件中填写的的Excel路径文件不存在')
        # if not is_sheet_exist(jasperSourcesExcelPath, sheetName):
        #     workbook = load_workbook(filename=jasperSourcesExcelPath)
        #     # 获取当前激活的sheet名称
        #     activeSheetName = workbook.active.title
        #     sheetName = activeSheetName
        #     print(f"配置文件中的填写的Sheet表不存在，默认使用当前激活的Sheet表是: {activeSheetName}")

        if not os.path.exists(jasperSourcesExcelPath):
            print('Excel路径文件不存在')
        workbook = load_workbook(filename=jasperSourcesExcelPath)
        # 获取当前激活的sheet名称
        activeSheetName = workbook.active.title
        sheetName = activeSheetName

        # 处理数据源Excel
        exportFolder = pdfDir + '\\'
        df = None
        if jasperModeInit == '拣货签':
            df = jianHuoJasperSourcesExcel(jasperSourcesExcelPath, sheetName)
        elif jasperModeInit == '外箱签':
            df = waiXiangJasperSourcesExcel(jasperSourcesExcelPath, sheetName)
        # 输出Json
        # outputJsonFilePath = 'D:\\stocktaking\\reports\\'
        mkDir2(outputJsonFilePath)
        jsonName = jasperMode + '.json'
        excelToJson(outputJsonFilePath + jsonName, df)
        jrxmlName = jasperMode + '.jrxml'
        pdfFileName = jasperModeInit + nowTime() + '.pdf'
        processingReportJapser(reportDir, jrxmlName, outputJsonFilePath + jsonName, exportFolder + pdfFileName)
        print('输出拣货签(PDF格式)' + exportFolder + pdfFileName + '完成')
    except Exception as e:
        print(f"错误信息: {e}")
        logging.error(nowTimeSecond())
        traceback.print_exc()  # 打印完整的堆栈跟踪信息
        # logging.error(f'错误信息: {e}')
        logging.error("An error occurred:", exc_info=True)
        input('回车退出程序')
